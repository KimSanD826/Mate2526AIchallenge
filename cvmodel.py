import cv2
import logging
from ultralytics import YOLO
import time
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os
from deepsort_tracker import DeepSORTTracker

def process_video(model_path, video_path, output_path=None, conf_threshold=0.5, plot_output_path=None):
    # Load model
    model = YOLO(model_path)
    
    # Open video
    cap = cv2.VideoCapture(video_path)
    assert cap.isOpened(), "Error opening video file"
    
    # Get video properties for output if needed
    fps = cap.get(cv2.CAP_PROP_FPS)
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    
    # Video writer if output path is specified
    if output_path:
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        out = cv2.VideoWriter(output_path, fourcc, fps, (width, height))
    
    # For FPS calculation
    prev_time = 0
   
    #  Setup for plotting fish counts every 5 seconds
    plt.ion()
    fig, ax = plt.subplots()
    sampled_times_seconds = []
    sampled_fish_counts = []
    sample_interval_seconds = 5.0
    next_sample_time_seconds = sample_interval_seconds
    
    # Setup Excel spreadsheet for fish count data
    wb = Workbook()
    ws = wb.active
    ws.title = "Fish Count Data"
    
    # Add headers with styling
    headers = ["Time (seconds)", "Fish Count", "Track IDs", "Timestamp"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    row_counter = 2  # Start from row 2 (after headers)
    
    # Initialize DeepSORT tracker
    tracker = DeepSORTTracker(max_age=30, min_hits=3, iou_threshold=0.3)
    
    while cap.isOpened():
        success, frame = cap.read()
        if not success:
            print("Video ended or frame read failed")
            break
        
        # Run inference
        results = model(frame, conf=conf_threshold)
        
        # Calculate FPS
        current_time = time.time()
        fps = 1 / (current_time - prev_time) if prev_time != 0 else 0
        prev_time = current_time
        
        # Extract detections for tracking
        detections = []
        if results and hasattr(results[0], "boxes") and results[0].boxes is not None:
            try:
                boxes = results[0].boxes
                if hasattr(boxes, 'xyxy'):
                    # Convert to numpy array if needed
                    if hasattr(boxes.xyxy, 'cpu'):
                        boxes_xyxy = boxes.xyxy.cpu().numpy()
                    else:
                        boxes_xyxy = boxes.xyxy
                    
                    for box in boxes_xyxy:
                        # Handle different box formats
                        if len(box) >= 4:
                            x1, y1, x2, y2 = box[:4]
                            # Ensure we have at least 6 values for tracking
                            conf = box[4] if len(box) > 4 else 0.5
                            cls = box[5] if len(box) > 5 else 0
                            detections.append([x1, y1, x2, y2, conf, cls])
                else:
                    # Fallback for older versions
                    for i in range(len(boxes)):
                        box = boxes[i]
                        if hasattr(box, 'xyxy'):
                            x1, y1, x2, y2 = box.xyxy[0]
                            conf = box.conf[0] if hasattr(box, 'conf') else 0.5
                            cls = box.cls[0] if hasattr(box, 'cls') else 0
                            detections.append([x1, y1, x2, y2, conf, cls])
            except Exception as e:
                print(f"Error processing detections: {e}")
                detections = []
        
        # Update DeepSORT tracker
        tracked_objects = tracker.update(detections, frame)
        
        # Count detections (fish) in this frame
        fish_count = len(detections)

        # Determine current video timestamp in seconds
        current_video_time_seconds = cap.get(cv2.CAP_PROP_POS_MSEC) / 1000.0

        # Sample and update the plot every 5 seconds of video time
        while current_video_time_seconds >= next_sample_time_seconds:
            sampled_times_seconds.append(next_sample_time_seconds)
            sampled_fish_counts.append(fish_count)
            
            # Add data to Excel spreadsheet
            current_timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
            
            # Get track IDs for current frame
            track_ids = []
            for obj in tracked_objects:
                # The tracker returns Track objects, not dictionaries
                if hasattr(obj, 'time_since_update') and obj.time_since_update == 0:
                    track_ids.append(str(obj.track_id))
            track_ids_str = ", ".join(track_ids) if track_ids else "None"
            
            ws.cell(row=row_counter, column=1, value=next_sample_time_seconds)
            ws.cell(row=row_counter, column=2, value=fish_count)
            ws.cell(row=row_counter, column=3, value=track_ids_str)
            ws.cell(row=row_counter, column=4, value=current_timestamp)
            row_counter += 1

            ax.clear()
            ax.plot(sampled_times_seconds, sampled_fish_counts, marker='o')
            ax.set_xlabel('Time (s)')
            ax.set_ylabel('Fish count')
            ax.set_title('Fish count every 5 seconds')
            ax.grid(True)
            plt.tight_layout()
            plt.pause(0.001)

            next_sample_time_seconds += sample_interval_seconds
        
        # Visualize results
        annotated_frame = results[0].plot()
        
        # Draw tracking information on the frame
        for obj in tracked_objects:
            if hasattr(obj, 'time_since_update') and obj.time_since_update == 0:  # Only draw active tracks
                bbox = obj.bbox
                track_id = obj.track_id
                age = obj.age
                
                # Draw bounding box with track ID
                x1, y1, x2, y2 = map(int, bbox)
                cv2.rectangle(annotated_frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                
                # Draw track ID and age
                label = f"ID:{track_id} Age:{age}"
                cv2.putText(annotated_frame, label, (x1, y1-10), 
                           cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
        
        # Display runtime FPS and current fish count
        cv2.putText(annotated_frame, f"FPS: {int(fps)}", (10, 30),
                   cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
        cv2.putText(annotated_frame, f"Fish: {fish_count}", (10, 60),
                   cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2)
        cv2.putText(annotated_frame, f"Tracks: {len(tracked_objects)}", (10, 90),
                   cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
        
        # Write to output file if specified
        if output_path:
            out.write(annotated_frame)
        
        # Display
        cv2.imshow("Fish Classification", annotated_frame)
        
        # Break the loop if 'q' is pressed or window is closed
        if cv2.waitKey(1) & 0xFF == ord("q"):
            break
        if cv2.getWindowProperty("Fish Classification", cv2.WND_PROP_VISIBLE) < 1:
            break
    
    # Clean up
    cap.release()
    if output_path:
        out.release()
    cv2.destroyAllWindows()
    try:
        plt.ioff()
        # Save the plot to a file for later viewing
        if plot_output_path is None:
            plot_output_path = "fish_counts.png"
        try:
            fig.savefig(plot_output_path, dpi=200)
            print(f"Saved fish count plot to {plot_output_path}")
        except Exception as e:
            print(f"Failed to save plot: {e}")
        
        # Save Excel spreadsheet
        excel_filename = "fish_count_data.xlsx"
        try:
            wb.save(excel_filename)
            print(f"Saved fish count data to Excel file: {excel_filename}")
        except Exception as e:
            logging.error(f"Failed to save Excel file: {e}")
        
        if len(sampled_times_seconds) > 0:
            plt.show()
    except Exception:
        logging.error("Failed to save plot")
    print("Video processing completed")

# convert info here
process_video(
    model_path="best.pt",  
    video_path="computer_coding_challenge_2025.mp4",
    output_path="output.mp4",  
    conf_threshold=0.6,
    plot_output_path="fish_counts.png"
)